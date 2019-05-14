import datetime
from django.conf import settings
import logging
from openpyxl import load_workbook
from collections import OrderedDict
import psycopg2
import pytz
import titlecase

PERTH = pytz.timezone('Australia/Perth')
LOGGER = logging.getLogger('sync_tasks')
alesco_date_to_dt = lambda d: PERTH.localize(datetime.datetime(d.year, d.month, d.day, 0, 0))+datetime.timedelta(hours=17, minutes=30)

def alesco_scrub_title(title):
    # remove extra spaces
    title_raw = ' '.join(title.upper().split())

    # ranger job titles have too much junk attached, jettison!
    if title_raw.startswith('RANGER'):
        return 'Ranger'
    if title_raw.startswith('SENIOR RANGER'):
        return 'Senior Ranger'


    def replace(word, **kwargs):
        result = None
        prefix = ''
        suffix = ''
        if word.startswith('('):
            prefix = '('
            word = word[1:]
        if word.endswith(')'):
            suffix = ')'
            word = word[:-1]
        if word.upper() in ('RIA', 'DBCA', 'BGPA', 'ZPA', 'PVS', 'IT', 'ICT', 'AV', 'HR', 'GIS', 'FMDP', 'SFM', 'OIM', 'HAAMC', 'TEC', 'MATES', 'AWU', 'FOI', 'KABC', 'VOG', 'WSS', 'EDRMS', 'LUP', 'WA', 'KSCS', 'OT'):
            result = word.upper()
        else:
            expand = {
                '&': 'and',
                'BG': 'Botanic Garden',
                'DEPT': 'Department',
                'CONS': 'Conservation',
                'CONSER': 'Conservation',
                'CONSERV': 'Conservation',
                'COORD': 'Coordinator',
                'CO-ORDINATOR': 'Coordinator',
                'COORDIN': 'Coordinator',
                'CUST': 'Customer',
                'MGMT': 'Management',
                'NP': 'National Park',
                'OCC': 'Occupational',
                'SAF': 'Safety',
                'SRV': 'Service',
                'SNR': 'Senior',
                'SERVIC': 'Services',
                'SCIENT': 'Scientist',
                'SCIENT.': 'Scientist',
                'ODG': 'Office of the Director General',
                'CHAIRPERSON,': 'Chairperson -',
                'OFFICER,': 'Officer -',
                'DIRECTOR,': 'Director -',
                'LEADER,': 'Leader -',
                'MANAGER,': 'Manager -',
                'COORDINATOR,': 'Coordinator -',
            }
            if word.upper() in expand:
                result = expand[word.upper()]
        if result:
            return prefix+result+suffix

    title_fixed = titlecase.titlecase(title_raw, callback=replace)
    return title_fixed


def alesco_data_import(fileobj):
    """Import task expects to be passed a file object (an uploaded .xlsx).
    """
    from .models import DepartmentUser
    LOGGER.info('Alesco data for DepartmentUsers is being updated')
    wb = load_workbook(fileobj, read_only=True)
    ws = wb.worksheets[0]
    keys = []
    values = []
    non_matched = 0
    multi_matched = 0
    updates = 0
    # Iterate over each row in the worksheet.
    for k, row in enumerate(ws.iter_rows()):
        values = []
        for cell in row:
            # First row: generate keys.
            if k == 0:
                keys.append(cell.value)
            # Otherwise make a list of values.
            else:
                # Serialise datetime objects.
                if isinstance(cell.value, datetime.datetime):
                    values.append(cell.value.isoformat())
                else:
                    values.append(cell.value)
        if k > 0:
            # Construct a dictionary of row values.
            record = dict(zip(keys, values))
            # Try to find a matching DepartmentUser by employee id.
            d = DepartmentUser.objects.filter(employee_id=record['EMPLOYEE_NO'])
            if d.count() > 1:
                multi_matched += 1
            elif d.count() == 1:
                d = d[0]
                d.alesco_data = record
                d.save()
                updates += 1
            else:
                non_matched += 0
    if updates > 0:
        LOGGER.info('Alesco data for {} DepartmentUsers was updated'.format(updates))
    if non_matched > 0:
        LOGGER.warning('Employee ID was not matched for {} rows'.format(non_matched))
    if multi_matched > 0:
        LOGGER.error('Employee ID was matched for >1 DepartmentUsers for {} rows'.format(multi_matched))
    return True


def update_user_from_alesco(user):
    from .models import DepartmentUser
    today = alesco_date_to_dt(datetime.date.today())
    term_date = None
    manager = None
    title = None
    changes = False
    if user.alesco_data:
        term_dates = [datetime.date.fromisoformat(x['job_term_date']) for x in user.alesco_data if x['job_term_date']]
        if term_dates:
            term_date = max(term_dates)
            term_date = alesco_date_to_dt(term_date) if term_date and term_date != ALESCO_DATE_MAX else None
        managers = [x['manager_emp_no'] for x in user.alesco_data if x['manager_emp_no']]
        managers = OrderedDict.fromkeys(managers).keys()
        managers = [DepartmentUser.objects.filter(employee_id=x).first() for x in managers]
        managers = [x for x in managers if x and (user.pk != x.pk)]
        if managers:
            manager = managers[0]
        title = next((x['occup_pos_title'] for x in user.alesco_data if 'occup_pos_title' in x and x['occup_pos_title']), None)
        if title:
            title = alesco_scrub_title(title)

    if term_date:
        stored_term_date = PERTH.normalize(user.date_hr_term) if user.date_hr_term else None
        if term_date != stored_term_date:
            
            if user.hr_auto_expiry:
                LOGGER.info('Updating expiry for {} from {} to {}'.format( user.email, stored_term_date, term_date ))
                user.expiry_date = term_date
            user.date_hr_term = term_date
            changes = True

    if manager:
        if manager != user.parent:
            LOGGER.info('Updating manager for {} from {} to {}'.format(user.email, user.parent.email if user.parent else None, manager.email if manager else None))
            user.parent = manager
            changes = True

    if title:
        if title != user.title:
            LOGGER.info('Updating title for {} from {} to {}'.format(user.email, user.title, title))
            user.title = title
            changes = True
    
    if changes:
        user.save()


ALESCO_DB_FIELDS = (
    'employee_id', 'surname', 'initials', 'first_name', 'second_name', 'gender',
    'date_of_birth', 'occup_type', 'current_commence', 'job_term_date',
    'occup_commence_date', 'occup_term_date', 'position_id', 'occup_pos_title',
    'clevel1_id', 'clevel1_desc', 'clevel5_id', 'clevel5_desc', 'award',
    'classification', 'step_id', 'emp_status', 'emp_stat_desc',
    'location', 'location_desc', 'paypoint', 'paypoint_desc', 'manager_emp_no',
)
ALESCO_DATE_MAX = datetime.date(2049, 12, 31)


def alesco_db_fetch():
    """Returns an iterator which fields rows from a database query until completed.
    """
    conn = psycopg2.connect(
        host=settings.ALESCO_DB_HOST,
        database=settings.ALESCO_DB_NAME,
        user=settings.ALESCO_DB_USERNAME,
        password=settings.ALESCO_DB_PASSWORD
    )
    cur = conn.cursor()

    query = "SELECT {} FROM {};".format(', '.join(ALESCO_DB_FIELDS), settings.ALESCO_DB_TABLE)
    cur.execute(query)
    while True:
        row = cur.fetchone()
        if row is None:
            break
        yield row


def alesco_db_import():
    from .models import DepartmentUser

    date_fields = ['date_of_birth', 'current_commence', 'job_term_date', 'occup_commence_date', 'occup_term_date']

    status_ranking = [
        'PFAS', 'PFA', 'PFT', 'CFA', 'CFT', 'NPAYF',
        'PPA', 'PPT', 'CPA', 'CPT', 'NPAYP',
        'CCFA', 'CAS', 'SEAS', 'TRAIN', 'NOPAY', 'NON',
    ]

    classification_ranking = [
        'CEO', 'CL3', 'CL2', 'CL1',
        'L9', 'L8', 'L7',
        'SCL6', 'L6',
        'SCL5', 'L5',
        'SCL4', 'S4', 'L4',
        'SCL3', 'S3', 'L3',
        'SCL2', 'R2', 'L2',
        'SCL1', 'R1', 'L12', 'L1',
    ]

    date_to_dt = lambda d: PERTH.localize(datetime.datetime(d.year, d.month, d.day, 0, 0)) + datetime.timedelta(days=1)
    records = {}
    alesco_iter = alesco_db_fetch()

    for row in alesco_iter:

        record = dict(zip(ALESCO_DB_FIELDS, row))
        eid = record['employee_id']

        if eid not in records:
            records[eid] = []
        records[eid].append(record)

    users = []

    for key, record in records.items():
        record.sort(key=lambda x: classification_ranking.index(x['classification']) if x['classification'] in classification_ranking else 100)
        record.sort(key=lambda x: status_ranking.index(x['emp_status']) if x['emp_status'] in status_ranking else 100)
        record.sort(key=lambda x: x['job_term_date'], reverse=True)

        for rec in record:
            for field in date_fields:
                rec[field] = rec[field].isoformat() if rec[field] and rec[field] != ALESCO_DATE_MAX else None

        user = DepartmentUser.objects.filter(employee_id=key).first()

        if not user:
            continue

        user.alesco_data = record

        user.save()
        update_user_from_alesco(user)
        users.append(user)

    return users


"""def alesco_list_expiry():
    from .models import DepartmentUser
    h_calmp = ALESCO_DB_FIELDS
    user_map = {}

    for row in alesco_db_fetch():
        term_date = PERTH.localize(datetime.datetime(row[9].year, row[9].month, row[9].day, 0, 0))
        if row[0] not in user_map:
            user_map[row[0]] = term_date
        elif term_date > user_map[row[0]]:
            user_map[row[0]] = term_date

    today = datetime.date.today()
    now = datetime.datetime.utcnow().replace(tzinfo=datetime.timezone.utc)

    end = datetime.datetime(2049, 1, 1, tzinfo=datetime.timezone.utc)

    results = []
    for user in DepartmentUser.objects.filter(active=True):
        if user.employee_id in user_map:
            test_time = user.expiry_date if user.expiry_date is not None else end
            if user_map[user.employee_id] < test_time:
                res = [user.email, user.given_name, user.surname, user.employee_id, user.ad_dn, user_map[user.employee_id], test_time]
                results.append(res)

    return results
    #outsiders  = [x for x in results if x[6] > now and x[5] < now]


    #with open('fixey.csv', 'w') as f:
    #    c = csv.writer(f)
    #    c.writerow(['email', 'given_name', 'surname', 'employee_id', 'ad_dn', 'alesco_expiry', 'itassets_expiry'])
    #    for r in results:
    #        c.writerow(r)
"""
