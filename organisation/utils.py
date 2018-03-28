from datetime import datetime, timedelta
from django.conf import settings
from django.core.files.base import ContentFile
import logging
from openpyxl import load_workbook
import os
from six import BytesIO
import subprocess
import unicodecsv


# Python 2 can't serialize unbound functions, so here's some dumb glue
def get_photo_path(instance, filename='photo.jpg'):
    return os.path.join('user_photo', '{0}.{1}'.format(instance.id, os.path.splitext(filename)))


def get_photo_ad_path(instance, filename='photo.jpg'):
    return os.path.join('user_photo_ad', '{0}.{1}'.format(instance.id, os.path.splitext(filename)))


def logger_setup(name):
    # Ensure that the logs dir is present.
    subprocess.check_call(['mkdir', '-p', 'logs'])
    # Set up logging in a standardised way.
    logger = logging.getLogger(name)
    if settings.DEBUG:
        logger.setLevel(logging.DEBUG)
    else:  # Log at a higher level when not in debug mode.
        logger.setLevel(logging.INFO)
    if not len(logger.handlers):  # Avoid creating duplicate handlers.
        fh = logging.handlers.RotatingFileHandler(
            'logs/{}.log'.format(name), maxBytes=5 * 1024 * 1024, backupCount=5)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    return logger


def alesco_data_import(filepath):
    """Import task expects to be passed a file path to a closed .xlsx file.
    """
    from .models import DepartmentUser
    logger = logger_setup('alesco_data_import')
    f = open(filepath)
    wb = load_workbook(filename=f.name, read_only=True)
    ws = wb.worksheets[0]
    keys = []
    values = []
    non_matched = 0
    multi_matched = 0
    updates = 0
    # Iterate over each row in the worksheet.
    for k, row in enumerate(ws.iter_rows()):
        values = []
        for cell in row:
            # First row: generate keys.
            if k == 0:
                keys.append(cell.value)
            # Otherwise make a list of values.
            else:
                # Serialise datetime objects.
                if isinstance(cell.value, datetime):
                    values.append(cell.value.isoformat())
                else:
                    values.append(cell.value)
        if k > 0:
            # Construct a dictionary of row values.
            record = dict(zip(keys, values))
            # Try to find a matching DepartmentUser by employee id.
            d = DepartmentUser.objects.filter(employee_id=record['EMPLOYEE_NO'])
            if d.count() > 1:
                multi_matched += 1
            elif d.count() == 1:
                d = d[0]
                d.alesco_data = record
                d.save()
                logger.info('Alesco data updated for {}'.format(d.email.lower()))
                updates += 1
            else:
                non_matched += 0
    if updates > 0:
        logger.info('Alesco data for {} DepartmentUsers was updated.'.format(updates))
    if non_matched > 0:
        logger.warning('Employee ID was not matched for {} rows.'.format(non_matched))
    if multi_matched > 0:
        logger.error('Employee ID was matched for >1 DepartmentUsers for {} rows.'.format(multi_matched))

    return True


def departmentuser_csv_report():
    """Output data from all DepartmentUser objects to a CSV, unpacking the
    various JSONField values.
    Returns a BytesIO object that can be written to a response or file.
    """
    from .models import DepartmentUser
    FIELDS = [
        'email', 'username', 'given_name', 'surname', 'name', 'preferred_name', 'title',
        'name_update_reference', 'employee_id', 'active', 'telephone', 'home_phone',
        'mobile_phone', 'other_phone', 'extension', 'expiry_date', 'org_unit',
        'cost_centre', 'parent', 'executive', 'vip', 'security_clearance',
        'in_sync', 'contractor', 'ad_deleted', 'o365_licence', 'shared_account',
        'populate_primary_group', 'notes', 'working_hours', 'sso_roles', 'org_data', 'alesco_data',
        'ad_data', 'extra_data', 'date_created', 'date_ad_updated', 'date_updated', 'ad_dn',
        'ad_guid']

    # Get any DepartmentUser with non-null alesco_data field.
    # alesco_data structure should be consistent to all (or null).
    du = DepartmentUser.objects.filter(alesco_data__isnull=False)[0]
    alesco_fields = du.alesco_data.keys()
    alesco_fields.sort()
    org_fields = {
        'department': ('units', 0, 'name'),
        'tier_2': ('units', 1, 'name'),
        'tier_3': ('units', 2, 'name'),
        'tier_4': ('units', 3, 'name'),
        'tier_5': ('units', 4, 'name')
    }

    header = [f for f in FIELDS]
    # These fields appended manually:
    header.append('account_type')
    header.append('position_type')
    header += org_fields.keys()
    header += alesco_fields

    # Get any DepartmentUser with non-null org_data field for the keys.
    if DepartmentUser.objects.filter(org_data__isnull=False).exists():
        du = DepartmentUser.objects.filter(org_data__isnull=False)[0]
        cc_keys = du.org_data['cost_centre'].keys()
        header += ['cost_centre_{}'.format(k) for k in cc_keys]
        location_keys = du.org_data['location'].keys()
        header += ['location_{}'.format(k) for k in location_keys]
        header.append('secondary_location')

    # Get any DepartmentUser with non-null ad_data field for the keys.
    if DepartmentUser.objects.filter(ad_data__isnull=False).exists():
        du = DepartmentUser.objects.filter(ad_data__isnull=False)[0]
        ad_keys = du.ad_data.keys()
        if 'mailbox' in ad_keys:
            ad_keys.remove('mailbox')  # Remove the nested object.
        header += ['ad_{}'.format(k) for k in ad_keys]

    # Write data for all DepartmentUser objects to the CSV
    stream = BytesIO()
    wr = unicodecsv.writer(stream, encoding='utf-8')
    wr.writerow(header)
    for u in DepartmentUser.objects.all():
        record = []
        for f in FIELDS:
            record.append(getattr(u, f))
        try:  # Append account_type display value.
            record.append(u.get_account_type_display())
        except:
            record.append('')
        try:  # Append position_type display value.
            record.append(u.get_position_type_display())
        except:
            record.append('')
        for o in org_fields:
            try:
                src = u.org_data
                for x in org_fields[o]:
                    src = src[x]
                record.append(src)
            except:
                record.append('')

        for a in alesco_fields:
            try:
                record.append(u.alesco_data[a])
            except:
                record.append('')
        for i in cc_keys:
            try:
                record.append(u.org_data['cost_centre'][i])
            except:
                record.append('')
        for i in location_keys:
            try:
                record.append(u.org_data['location'][i])
            except:
                record.append('')
        if u.org_data and 'secondary_location' in u.org_data:
            record.append(u.org_data['secondary_location'])
        else:
            record.append('')
        for i in ad_keys:
            try:
                record.append(u.ad_data[i])
            except:
                record.append('')

        # Write the row to the CSV stream.
        wr.writerow(record)

    return stream.getvalue()


def convert_ad_timestamp(timestamp):
    """Converts an Active Directory timestamp to a Python datetime object.
    Ref: http://timestamp.ooz.ie/p/time-in-python.html
    """
    epoch_start = datetime(year=1601, month=1, day=1)
    seconds_since_epoch = timestamp / 10**7
    return epoch_start + timedelta(seconds=seconds_since_epoch)


def load_mugshots(data_dir='/root/mugshots'):
    from .models import DepartmentUser
    files = [x for x in os.listdir(data_dir) if os.path.isfile(os.path.join(data_dir, x))]
    valid = 0
    for f in files:
        name = os.path.splitext(f)[0]
        qs = DepartmentUser.objects.filter(username__iexact=name)
        if qs:
            with open(os.path.join(data_dir, f)) as fp:
                qs[0].photo.save(f, ContentFile(fp.read()))
            print('Updated photo for {}'.format(name))
            valid += 1
        else:
            print('ERROR: Username {} not found'.format(name))

    print('{}/{} photos valid'.format(valid, len(files)))
